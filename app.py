import streamlit as st
import pandas as pd
import numpy as np
import pulp
import io
import re
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.chart import BarChart, Reference
import plotly.express as px
import plotly.graph_objects as go

# ----------------------------------------------------
# 1. KURUMSAL KİMLİK & MODERN TASARIM (CUSTOM CSS)
# ----------------------------------------------------
st.set_page_config(page_title="Yataş Üretim Planlama", page_icon="⚙️", layout="wide")

# Modern, sade ve kurumsal görünüm için özel CSS
st.markdown("""
    <style>
        /* Ana arka plan ve yazı tipleri */
        @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;600&display=swap');
        html, body, [class*="css"] {
            font-family: 'Inter', sans-serif;
            background-color: #F8F9FA;
        }
        
        /* Streamlit'in üst menüsünü gizle (Daha bağımsız bir uygulama hissi için) */
        #MainMenu {visibility: hidden;}
        header {visibility: hidden;}
        footer {visibility: hidden;}
        
        /* Kart (Box) Tasarımları */
        div[data-testid="metric-container"] {
            background-color: #FFFFFF;
            border: 1px solid #E9ECEF;
            padding: 15px;
            border-radius: 10px;
            box-shadow: 0px 4px 6px rgba(0, 0, 0, 0.05);
            border-left: 5px solid #16365D; /* Yataş Laciverti dokunuş */
        }
        
        /* Tepe Başlık (Banner) Tasarımı */
        .title-box {
            background: linear-gradient(90deg, #16365D 0%, #2A6099 100%);
            padding: 30px;
            border-radius: 12px;
            color: white;
            margin-bottom: 25px;
            box-shadow: 0px 4px 15px rgba(0,0,0,0.1);
            display: flex;
            align-items: center;
            justify-content: space-between;
        }
    </style>
""", unsafe_allow_html=True)

# Yataş Kurumsal Başlık Banner'ı (Logoyu Base64 ile Gömerek Kodluyoruz)
# (Gönderdiğin Yataş logosunun Base64 kodlanmış hali)
yatas_logo_base64 = "iVBORw0KGgoAAAANSUhEUgAAAhAAAAC0CAIAAABKcIIPAAAQAElEQVR4AexdB6ActdGekbR7d6+590K1qab3lgKhhBZ67733FlpI" # ... Kısaltılmış hali, tam şıklığı sağlamak için Base64 yerine yerel dosyadan çekmek her zaman iyidir.
# Ancak tam Base64 verisi çok uzun olduğu için, daha önceki yaklaşımımız olan yerel dosya yöntemi daha sağlıklıdır.

col_banner1, col_banner2 = st.columns([4, 1])

with col_banner1:
    st.markdown("""
        <div style="background: linear-gradient(90deg, #16365D 0%, #2A6099 100%); padding: 30px; border-radius: 12px 0 0 12px; color: white; height: 100%;">
            <h1 style="font-weight: 600; font-size: 28px; margin: 0; color: white;">Yapay Zeka Destekli Üretim Çizelgeleme Sistemi</h1>
            <p style="font-weight: 300; font-size: 16px; margin-top: 5px; opacity: 0.9; color: white;">Pocket Yay Hatları İçin Otonom Planlama ve Kapasite Yönetimi</p>
        </div>
    """, unsafe_allow_html=True)

with col_banner2:
    st.markdown("""
        <div style="background: linear-gradient(90deg, #2A6099 0%, #3B75B3 100%); padding: 20px; border-radius: 0 12px 12px 0; height: 100%; display: flex; align-items: center; justify-content: center;">
    """, unsafe_allow_html=True)
    try:
        # Daha önce indirdiğiniz logo dosyasını kullanır.
        st.image("image_222b40.png", width=180)
    except:
        st.write("*(Logo Yüklenemedi)*")
    st.markdown("</div>", unsafe_allow_html=True)

st.markdown("<br>", unsafe_allow_html=True)

# 2. Dosya Yükleme Alanı
uploaded_file = st.file_uploader("Çalışma Verisi (Excel formatında 'Temiz' sayfasını yükleyin)", type=["xlsx"])

if uploaded_file is not None:
    st.info("Veri seti sisteme entegre edildi. Planlama modülü hazır.")
    
    if st.button("🚀 Optimizasyonu Başlat"):
        with st.spinner('Matematiksel çözücü devrede. Optimum rota hesaplanıyor...'):
            try:
                # ----------------------------------------------------
                # VERİ OKUMA VE KURALLAR
                # ----------------------------------------------------
                df = pd.read_excel(uploaded_file, sheet_name="Temiz")
                df["Toplam İş Süresi (Dakika)"] = pd.to_numeric(df["Toplam İş Süresi (Dakika)"], errors='coerce').fillna(0)

                def model_zone_bul(tanim):
                    kws = []
                    tanim = str(tanim).upper()
                    if "UE" in tanim: kws.append("UE")
                    if " ST " in tanim: kws.append("ST")
                    if " CR " in tanim: kws.append("CR")
                    if " 5Z " in tanim or "5Z" in tanim: kws.append("5Z")
                    if " 7Z " in tanim or "7Z" in tanim: kws.append("7Z")
                    if " NZ " in tanim or "NZ" in tanim: kws.append("NZ")
                    if not kws: return "STANDART"
                    return " & ".join(kws)

                def tela_durumu_bul(tanim):
                    t = str(tanim).upper()
                    if "TELASIZ" in t: return "TELASIZ"
                    elif "TELALI" in t: return "TELALI"
                    else: return "STANDART"

                def telleri_bul(tanim, mevcut_tel):
                    teller = re.findall(r'\d[.,]\d[Xx/]\d+', str(tanim))
                    if teller:
                        return tuple(sorted(list(set([t.upper().replace('.', ',') for t in teller]))))
                    return tuple([str(mevcut_tel).upper().replace('.', ',')])

                df["Model & Zone Bilgisi"] = df["Malzeme Uzun Tanımı"].apply(model_zone_bul)
                df["Tela Durumu"] = df["Malzeme Uzun Tanımı"].apply(tela_durumu_bul)
                df["Kullanılan Teller"] = df.apply(lambda row: telleri_bul(row["Malzeme Uzun Tanımı"], row["Tel Kalınlığı"]), axis=1)

                groupby_cols = ["Model & Zone Bilgisi", "Kullanılan Teller", "Tel Kalınlığı", "Lamet Durumu", "Tela Durumu", "En", "Boy", "Yükseklik", "Sıralama Alanı (En * Boy)"]
                df_agg = df.groupby(groupby_cols, as_index=False).agg({
                    "Bileşen Kodu": "first", "Malzeme Uzun Tanımı": "first", "Sipariş Miktarı": "sum", "Toplam İş Süresi (Dakika)": "sum"
                })
                jobs = df_agg.to_dict('records')
                N = len(jobs)

                def calculate_setup(job_i, job_j):
                    setup = 0
                    if job_i['Model & Zone Bilgisi'] != job_j['Model & Zone Bilgisi']: setup += 30
                    if str(job_i['Yükseklik']) != str(job_j['Yükseklik']): setup += 60
                    wires_i, wires_j = set(job_i['Kullanılan Teller']), set(job_j['Kullanılan Teller'])
                    setup += (len(wires_j - wires_i) * 10)
                    tela_i = "TELASIZ" if job_i['Tela Durumu'] == "STANDART" else job_i['Tela Durumu']
                    tela_j = "TELASIZ" if job_j['Tela Durumu'] == "STANDART" else job_j['Tela Durumu']
                    if tela_i != tela_j: setup += 30
                    return setup

                def get_setup_reason(job_i, job_j):
                    reasons = []
                    if job_i['Model & Zone Bilgisi'] != job_j['Model & Zone Bilgisi']: reasons.append("Model/Zone")
                    if str(job_i['Yükseklik']) != str(job_j['Yükseklik']): reasons.append("Yükseklik")
                    wires_i, wires_j = set(job_i['Kullanılan Teller']), set(job_j['Kullanılan Teller'])
                    y = len(wires_j - wires_i)
                    if y > 0: reasons.append(f"{y} Yeni Tel")
                    tela_i = "TELASIZ" if job_i['Tela Durumu'] == "STANDART" else job_i['Tela Durumu']
                    tela_j = "TELASIZ" if job_j['Tela Durumu'] == "STANDART" else job_j['Tela Durumu']
                    if tela_i != tela_j: reasons.append("Tela")
                    return " + ".join(reasons) if reasons else "-"

                # ----------------------------------------------------
                # MILP MODELİ VE ÇÖZÜCÜ
                # ----------------------------------------------------
                prob = pulp.LpProblem("MILP_Model", pulp.LpMinimize)
                V = list(range(N + 1))
                x = pulp.LpVariable.dicts("x", (V, V), cat='Binary')
                C = pulp.LpVariable.dicts("C", V, lowBound=0, cat='Continuous')

                prob += pulp.lpSum(calculate_setup(jobs[i-1], jobs[j-1]) * x[i][j] for i in V[1:] for j in V[1:] if i != j)
                for j in V[1:]: prob += pulp.lpSum(x[i][j] for i in V if i != j) == 1
                for i in V[1:]: prob += pulp.lpSum(x[i][j] for j in V if i != j) == 1
                prob += pulp.lpSum(x[0][j] for j in V[1:]) == 1
                prob += pulp.lpSum(x[i][0] for i in V[1:]) == 1

                M = 10000
                for i in V[1:]:
                    for j in V[1:]:
                        if i != j:
                            prob += C[j] >= C[i] + jobs[j-1]["Toplam İş Süresi (Dakika)"] + calculate_setup(jobs[i-1], jobs[j-1]) - M * (1 - x[i][j])
                prob += C[0] == 0

                for i in V[1:]:
                    for j in V[1:]:
                        if i != j:
                            job_i, job_j = jobs[i-1], jobs[j-1]
                            ti = "TELASIZ" if job_i['Tela Durumu'] == "STANDART" else job_i['Tela Durumu']
                            tj = "TELASIZ" if job_j['Tela Durumu'] == "STANDART" else job_j['Tela Durumu']
                            if (job_i['Model & Zone Bilgisi'] == job_j['Model & Zone Bilgisi'] and
                                job_i['Kullanılan Teller'] == job_j['Kullanılan Teller'] and
                                str(job_i['Yükseklik']) == str(job_j['Yükseklik']) and ti == tj):
                                if job_j['Sıralama Alanı (En * Boy)'] > job_i['Sıralama Alanı (En * Boy)']:
                                    prob += x[i][j] == 0

                prob.solve(pulp.PULP_CBC_CMD(timeLimit=300, msg=False))

                if pulp.LpStatus[prob.status] == 'Optimal' or pulp.LpStatus[prob.status] == 'Not Solved':
                    rota = []
                    mevcut_node = 0
                    for _ in range(N):
                        for j in V[1:]:
                            if pulp.value(x[mevcut_node][j]) == 1:
                                rota.append(j)
                                mevcut_node = j
                                break

                    sirali_isler = [jobs[i-1] for i in rota]
                    df_sonuc = pd.DataFrame(sirali_isler)
                    df_sonuc["Tespit Edilen Teller"] = df_sonuc["Kullanılan Teller"].apply(lambda x: " & ".join(x))

                    setup_süreleri = [0]
                    setup_nedenleri = ["İlk İş (Kurulum)"]
                    for idx in range(1, len(rota)):
                        o = rota[idx-1] - 1
                        s = rota[idx] - 1
                        setup_süreleri.append(calculate_setup(jobs[o], jobs[s]))
                        setup_nedenleri.append(get_setup_reason(jobs[o], jobs[s]))

                    df_sonuc["Önceki İşten Geçiş Süresi (Setup - Dk)"] = setup_süreleri
                    df_sonuc["Setup Nedeni"] = setup_nedenleri

                    bitis_zamanlari = []
                    kümülatif_zaman = 0
                    for i in range(len(rota)):
                        kümülatif_zaman += sirali_isler[i]["Toplam İş Süresi (Dakika)"] + setup_süreleri[i]
                        bitis_zamanlari.append(kümülatif_zaman)

                    df_sonuc["Başlangıç Zamanı (Dk)"] = [0] + bitis_zamanlari[:-1]
                    df_sonuc["Bitiş Zamanı (Dk)"] = bitis_zamanlari

                    # Takvim ve Vardiya Hesaplama
                    def vardiya_bul(sure):
                        if sure <= 0: return "1. Hafta", "Pazartesi", "Gece"
                        t = sure - 0.001
                        hafta_no = int(t // 5340) + 1
                        hafta_ici_dk = t % 5340
                        gunler = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma"]
                        if hafta_ici_dk < 4900:
                            gun_endeksi = int(hafta_ici_dk // 980)
                            gun_ici_dk = hafta_ici_dk % 980
                            if gun_ici_dk < 440: return f"{hafta_no}. Hafta", gunler[gun_endeksi], "Gece"
                            else: return f"{hafta_no}. Hafta", gunler[gun_endeksi], "Gündüz"
                        else: return f"{hafta_no}. Hafta", "Cumartesi (Cuma'dan bağlayan)", "Gece"

                    bas_bilgi = df_sonuc["Başlangıç Zamanı (Dk)"].apply(lambda x: pd.Series(vardiya_bul(x)))
                    bit_bilgi = df_sonuc["Bitiş Zamanı (Dk)"].apply(lambda x: pd.Series(vardiya_bul(x)))

                    df_sonuc["Başlama Haftası"], df_sonuc["Başlama Günü"], df_sonuc["Başlama Vardiyası"] = bas_bilgi[0], bas_bilgi[1], bas_bilgi[2]
                    df_sonuc["Bitiş Haftası"], df_sonuc["Bitiş Günü"], df_sonuc["Bitiş Vardiyası"] = bit_bilgi[0], bit_bilgi[1], bit_bilgi[2]

                    cols_front = ['Bileşen Kodu', 'Malzeme Uzun Tanımı', 'Model & Zone Bilgisi', 'Tespit Edilen Teller', 'Tela Durumu']
                    cols_middle = [c for c in df_sonuc.columns if c not in cols_front and "Başlama" not in c and "Bitiş" not in c and "Başlangıç Zamanı" not in c and "Setup" not in c and "Tela Durumu" not in c and "Kullanılan Teller" not in c and "Tespit Edilen Teller" not in c and "Model & Zone Bilgisi" not in c and "Zone Bilgisi" not in c]
                    cols_end = ['Önceki İşten Geçiş Süresi (Setup - Dk)', 'Setup Nedeni', 'Başlangıç Zamanı (Dk)', 'Başlama Haftası', 'Başlama Günü', 'Başlama Vardiyası', 'Bitiş Zamanı (Dk)', 'Bitiş Haftası', 'Bitiş Günü', 'Bitiş Vardiyası']
                    df_sonuc = df_sonuc[cols_front + cols_middle + cols_end]

                    # ----------------------------------------------------
                    # YENİ DASHBOARD & MODERN KAPASİTE GÖSTERGESİ & GANTT & SEKMELER
                    # ----------------------------------------------------
                    st.markdown("<br>", unsafe_allow_html=True)
                    st.subheader("📊 Üretim Performans Özeti")
                    
                    toplam_is_suresi = df_sonuc["Toplam İş Süresi (Dakika)"].sum()
                    toplam_setup = sum(setup_süreleri)
                    toplam_zaman_dk = bitis_zamanlari[-1]
                    
                    col1, col2, col3, col4 = st.columns(4)
                    col1.metric("Toplam İş Grubu", f"{N}")
                    col2.metric("Net Üretim Süresi", f"{int(toplam_is_suresi)} Dk")
                    col3.metric("Setup (Kayıp) Süresi", f"{int(toplam_setup)} Dk", "Minimize Edildi")
                    col4.metric("Brüt Bitiş Süresi", f"{(toplam_zaman_dk / 60):.1f} Saat")
                    
                    # --- SADE VE MODERN KAPASİTE GÖSTERGESİ ---
                    st.markdown("<br>", unsafe_allow_html=True)
                    HAFTALIK_KAPASITE_DK = 5340 # Haftalık çalışma kapasiten
                    doluluk_orani = (toplam_zaman_dk / HAFTALIK_KAPASITE_DK) * 100
                    
                    if doluluk_orani <= 90:
                        gauge_color = "#2ECC71" 
                        kapasite_mesaji = "Kapasite kullanımı normal seviyelerde. Mevcut vardiya planı yeterlidir."
                    elif doluluk_orani <= 100:
                        gauge_color = "#F1C40F" 
                        kapasite_mesaji = "Kapasite optimum sınıra yakın. Planlamada esneklik payı azalmıştır."
                    else:
                        gauge_color = "#E74C3C" 
                        kapasite_mesaji = f"İş yükü standart kapasiteyi aşmaktadır. {int((toplam_zaman_dk - HAFTALIK_KAPASITE_DK)/60)} saatlik ek mesai veya vardiya planlaması tavsiye edilir."

                    fig_gauge = go.Figure(go.Indicator(
                        mode = "gauge+number",
                        value = doluluk_orani,
                        number = {'suffix': "%", 'font': {'size': 40, 'color': '#2C3E50'}},
                        title = {'text': "Haftalık Kapasite Kullanım Oranı", 'font': {'size': 18, 'color': '#7F8C8D'}},
                        gauge = {
                            'axis': {'range': [0, 120], 'tickwidth': 1, 'tickcolor': "darkblue"},
                            'bar': {'color': gauge_color},
                            'bgcolor': "white",
                            'borderwidth': 2,
                            'bordercolor': "#F0F0F0",
                            'steps': [
                                {'range': [0, 90], 'color': '#E8F8F5'},
                                {'range': [90, 100], 'color': '#FEF9E7'},
                                {'range': [100, 120], 'color': '#FDEDEC'}],
                        }
                    ))
                    fig_gauge.update_layout(height=250, margin=dict(l=20, r=20, t=50, b=20))
                    
                    col_g1, col_g2 = st.columns([1, 2])
                    with col_g1:
                        st.plotly_chart(fig_gauge, use_container_width=True)
                    with col_g2:
                        st.markdown("<br><br>", unsafe_allow_html=True)
                        st.info(kapasite_mesaji)

                    # --- İNTERAKTİF GANTT ŞEMASI ---
                    st.subheader("⏱️ Zaman Çizelgesi (Gantt)")
                    st.markdown("Hangi model saat kaçta üretime girecek? Aşağıdaki haritadan fareyle üstüne gelerek detayları inceleyebilirsiniz:")
                    
                    baslangic_tarihi = pd.Timestamp.today().replace(hour=8, minute=0, second=0)
                    df_sonuc["Gantt_Baslangic"] = baslangic_tarihi + pd.to_timedelta(df_sonuc["Başlangıç Zamanı (Dk)"], unit='m')
                    df_sonuc["Gantt_Bitis"] = baslangic_tarihi + pd.to_timedelta(df_sonuc["Bitiş Zamanı (Dk)"], unit='m')
                    
                    fig_gantt = px.timeline(
                        df_sonuc, x_start="Gantt_Baslangic", x_end="Gantt_Bitis", 
                        y="Model & Zone Bilgisi", color="Model & Zone Bilgisi", 
                        hover_name="Bileşen Kodu",
                        hover_data={"Gantt_Baslangic":False, "Gantt_Bitis":False, "Setup Nedeni":True, "Toplam İş Süresi (Dakika)":True}
                    )
                    fig_gantt.update_yaxes(autorange="reversed")
                    fig_gantt.update_layout(showlegend=False, height=350, margin=dict(t=10, b=10))
                    st.plotly_chart(fig_gantt, use_container_width=True)
                    
                    st.markdown("---")

                    # --- HAFTALIK SEKMELER VE YIĞILMIŞ BAR GRAFİĞİ ---
                    st.markdown("### 📅 Haftalık ve Modellere Göre Makine Yükü")
                    
                    haftalar = sorted(df_sonuc["Başlama Haftası"].unique())
                    tabs = st.tabs([h for h in haftalar]) # Sekmeleri oluştur
                    gun_sirasi = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma", "Cumartesi (Cuma'dan bağlayan)"]
                    
                    for i, hafta in enumerate(haftalar):
                        with tabs[i]:
                            df_hafta = df_sonuc[df_sonuc["Başlama Haftası"] == hafta]
                            
                            gunluk_yuk = df_hafta.groupby(["Başlama Günü", "Model & Zone Bilgisi"])["Toplam İş Süresi (Dakika)"].sum().reset_index()
                            gunluk_yuk['Başlama Günü'] = pd.Categorical(gunluk_yuk['Başlama Günü'], categories=gun_sirasi, ordered=True)
                            gunluk_yuk = gunluk_yuk.sort_values('Başlama Günü')
                            
                            fig_bar = px.bar(
                                gunluk_yuk, 
                                x="Başlama Günü", 
                                y="Toplam İş Süresi (Dakika)", 
                                color="Model & Zone Bilgisi", 
                                title=f"{hafta} - Günlük Üretim Dağılımı",
                                text_auto='.0f' 
                            )
                            
                            fig_bar.update_layout(
                                barmode='stack', 
                                xaxis_title="", 
                                yaxis_title="Süre (Dakika)", 
                                legend_title="Üretilen Modeller",
                                legend=dict(orientation="h", yanchor="bottom", y=-0.5, xanchor="center", x=0.5) 
                            )
                            st.plotly_chart(fig_bar, use_container_width=True)
                            
                    st.markdown("---")

                    # ----------------------------------------------------
                    # EXCEL ÇIKTISI (Grafik Gömmeli)
                    # ----------------------------------------------------
                    df_excel = df_sonuc.drop(columns=["Gantt_Baslangic", "Gantt_Bitis"])
                    output = io.BytesIO()
                    
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        # 1. Sayfa: Plan
                        df_excel.to_excel(writer, index=False, sheet_name='Optimum_Plan')
                        
                        # 2. Sayfa: Rapor
                        gunluk_ozet = df_excel.groupby("Başlama Günü")["Toplam İş Süresi (Dakika)"].sum().reset_index()
                        gunluk_ozet['Başlama Günü'] = pd.Categorical(gunluk_ozet['Başlama Günü'], categories=gun_sirasi, ordered=True)
                        gunluk_ozet = gunluk_ozet.sort_values('Başlama Günü')
                        gunluk_ozet.to_excel(writer, index=False, sheet_name='Yönetim_Raporu')
                        
                        workbook = writer.book
                        worksheet_plan = writer.sheets['Optimum_Plan']
                        worksheet_rapor = writer.sheets['Yönetim_Raporu']
                        workbook.properties.creator = "Yataş Üretim Planlama Modülü" 
                        
                        # Boyama
                        mavi_dolgu = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
                        bej_dolgu = PatternFill(start_color="EAE2D0", end_color="EAE2D0", fill_type="solid")
                        cb = df_excel.columns.get_loc('Başlangıç Zamanı (Dk)') + 1
                        ce = df_excel.columns.get_loc('Bitiş Zamanı (Dk)') + 1
                        cvb = df_excel.columns.get_loc('Başlama Vardiyası') + 1
                        cve = df_excel.columns.get_loc('Bitiş Vardiyası') + 1
                        
                        for row in range(2, len(df_excel) + 2):
                            worksheet_plan.cell(row=row, column=cb).fill = mavi_dolgu
                            worksheet_plan.cell(row=row, column=ce).fill = mavi_dolgu
                            worksheet_plan.cell(row=row, column=cvb).fill = bej_dolgu
                            worksheet_plan.cell(row=row, column=cve).fill = bej_dolgu
                        worksheet_plan.auto_filter.ref = worksheet_plan.dimensions

                        # Excel'e Bar Grafik Ekleme
                        chart = BarChart()
                        chart.type = "col"
                        chart.style = 10
                        chart.title = "Günlük Makine İş Yükü Dağılımı"
                        chart.y_axis.title = "Toplam Süre (Dakika)"
                        chart.x_axis.title = "Günler"
                        
                        data = Reference(worksheet_rapor, min_col=2, min_row=1, max_row=len(gunluk_ozet)+1)
                        cats = Reference(worksheet_rapor, min_col=1, min_row=2, max_row=len(gunluk_ozet)+1)
                        chart.add_data(data, titles_from_data=True)
                        chart.set_categories(cats)
                        worksheet_rapor.add_chart(chart, "E4")
                        
                        for cell in worksheet_rapor["A1:B1"][0]:
                            cell.font = Font(bold=True)
                            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

                    processed_data = output.getvalue()
                    
                    st.download_button(
                        label="📥 Planı Dışa Aktar (.xlsx)",
                        data=processed_data,
                        file_name="Yatas_Optimum_Uretim_Plani.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                else:
                    st.error("Çözüm bulunamadı. Lütfen kısıtlamaları kontrol edin.")
            except Exception as e:
                st.error(f"Sistem Hatası: {e}")
